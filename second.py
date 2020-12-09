from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.formula.translate import Translator

wb = load_workbook('test.xlsx')
main = wb['main']
maxrow = len(main['A']) + 1


season = input('What season are you in? \n')


if season in wb.sheetnames:
    print('Match found')
    ws = wb[season]

else:
    if input('No matches found, create new sheet? (y/n) \n') == 'y':
        ws = wb.create_sheet(season)
        
        for name in range(1, maxrow):
            ws.cell(row=name,column=1).value = main.cell(row=name,column=1).value
        names = ws.cell(row=1, column=1)
        names.font = Font(bold=True)
        names.value = 'Names'
        total = ws.cell(row=1, column=2)
        total.font = Font(bold=True)
        total.value = 'Total'
        print('New season: ' + str(season) + ' created successfully!')
    else:
        print('ok')

maxcol = ws.max_column
nextcol = maxcol + 1


for name in range(2, maxrow):
    #ask user how many goals they scored this week
    goals = input('How many goals did ' + str(ws.cell(row=name,column=1).value) + ' score this week? \n')
    #create new cell at len(ws)+1
    newcell = ws.cell(row=name, column=nextcol)
    newcell.font = Font(bold=False)
    newcell.value = int(goals)
    week = ws.cell(row=1, column=nextcol)
    week.font = Font(bold=True)
    week.value = 'Week ' + str(nextcol - 2)

for name in range(2, maxrow):
    newcell = ws.cell(row=name, column=2)
    newcell.font = Font(bold=False)
    newcell.value = "=SUM(C" + str(name) + ":ZZ" + str(name) + ")"


nextcolmain = main.max_column + 1
for sheet in wb:
    if sheet.title != 'main':
        
        season = main.cell(row=1, column=nextcolmain)
        season.font = Font(bold=True)
        season.value = sheet.title


for name in range(2, maxrow):
    sumtotal = "=SUM("
    for sheet in wb:
        if sheet.title != 'main':
            sumtotal += "+" + str(sheet.title) + "!" + "B" + str(name)
    sumtotal = sumtotal[0:5] + sumtotal[6:]
    sumtotal += ")"
    print(sumtotal)
    newcell = main.cell(row=name, column=2)
    newcell.font = Font(bold=False)
    newcell.value = sumtotal


wb.save('test.xlsx')
    