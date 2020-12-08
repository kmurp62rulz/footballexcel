from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

wb = Workbook()
main = wb.create_sheet('main')

names = main.cell(row=1, column=1)
names.font = Font(bold=True)
names.value = 'Names'

total = main.cell(row=1, column=2)
total.font = Font(bold=True)
total.value = 'Total'

numberOfPlayers = int(input('How many players on the team? \n'))
x = 2
while numberOfPlayers > 0:
    
    d = main.cell(row=x, column=1)
    d.value = input('Player name: \n')
    x += 1
    numberOfPlayers -= 1

sht1 = wb['Sheet']
wb.remove(sht1)

wb.save('test.xlsx')